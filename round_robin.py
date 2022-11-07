from openpyxl import load_workbook

processes = []
queue = []
d = {}
waiting_precesses = 0
worksheet = load_workbook(filename="cpu-scheduling.xlsx")
ws = worksheet.active

for i in ws.iter_rows(values_only=True):
    try:
        if type(i[0]) == str:
            continue
        else:
            processes.append(list(i))
    except ValueError:
        print(ValueError)


class Processes:
    def __init__(self, pid, arrival_time, burst_time):
        self.pid = pid
        self.arrival_time = arrival_time
        self.burst_time = burst_time


def round_robin(processes_list, num_of_processes):
    time_unit = 1
    waiting_precesses = 0
    finished_processes = 0
    time_quantum = 4

    for process in processes_list:
        if time_unit >= process.arrival_time:
            def check_new_arrival():
                global waiting_precesses
                for i in range(waiting_precesses, num_of_processes):
                    if time_unit >= processes_list[i].arrival_time:
                        new_arrived_process = processes_list[i]
                        queue.append(new_arrived_process.pid)
                        d[new_arrived_process.pid] = 0
                        waiting_precesses += 1

            def check_stat():
                x = []
                for i in queue:
                    d[i] += 1
                    x.append(f"PID: {i} Wait={d[i]}TU")
                return " ".join(map(str, x))

            if process.burst_time > time_quantum:
                check_new_arrival()
                queue.remove(process.pid)
                for j in range(time_quantum):
                    process.burst_time -= 1
                    instructions = process.burst_time
                    check_new_arrival()
                    if len(queue) < 1:
                        if instructions != 0:
                            print(
                                f"Time Unit: {time_unit} PID: {process.pid} executes. "
                                f"{instructions} instructions left. Q={time_quantum - j - 1} No queue")
                        elif instructions == 0:
                            print(
                                f"Time Unit: {time_unit} PID: {process.pid} "
                                f"last instruction. CPU is idle.")
                            print("All processes have been executed. End of simulation.")
                    else:
                        if instructions != 0:
                            print(
                                f"Time Unit: {time_unit} PID: {process.pid} executes. "
                                f"{instructions} instructions left Q={time_quantum - j - 1} Queue={len(queue)} {check_stat()}")
                        elif instructions == 0:
                            print(
                                f"Time Unit: {time_unit} PID: {process.pid} "
                                f"last instruction. Q={time_quantum - j - 1}")
                    time_unit += 1
                print(f"Time Unit: {time_unit} Context switch.{check_stat()}")
                time_unit += 1
                queue.append(process.pid)
                # d[process.pid] = 0
                processes_list.append(process)

            else:
                check_new_arrival()
                if process.pid in queue:
                    queue.remove(process.pid)
                for k in range(process.burst_time):
                    process.burst_time -= 1
                    instructions = process.burst_time
                    check_new_arrival()
                    if len(queue) < 1:
                        if instructions != 0:
                            print(
                                f"Time Unit: {time_unit} PID: {process.pid} executes. "
                                f"{instructions} instructions left. No queue")
                        elif instructions == 0:
                            print(
                                f"Time Unit: {time_unit} PID: {process.pid} "
                                f"last instruction. CPU is idle.")
                            print("All processes have executed. End of simulation.")
                    else:
                        if instructions != 0:
                            print(
                                f"Time Unit: {time_unit} PID: {process.pid} executes. "
                                f"{instructions} instructions left. Q={time_quantum - k - 1} Queue={len(queue)} {check_stat()}")
                        elif instructions == 0:
                            print(
                                f"Time Unit: {time_unit} PID: {process.pid} "
                                f"last instruction. Q={time_quantum - k - 1} {check_stat()}")
                    time_unit += 1
                    if process.burst_time == 0:
                        finished_processes += 1
                        waiting_precesses -= 1
                if len(queue) < 1 and process.burst_time == 0:
                    continue
                else:
                    print(f"Time Unit: {time_unit} Context switch.{check_stat()}")
                    time_unit += 1
        else:
            processes_list.append(process)


p = []
for i in processes:
    p.append(Processes(i[0], i[1], i[2]))

c = sorted(p, key=lambda x: (x.arrival_time, x.burst_time))
round_robin(c, len(c))
