from openpyxl import load_workbook

processes = []
queue = []
d = {}
waiting_precesses = 0
worksheet = load_workbook(filename="cpu-scheduling.xlsx")
ws = worksheet.active

for i in ws.iter_rows(values_only=True):
    try:
        if type(i[0]) == str:
            continue
        else:
            processes.append(list(i))
    except ValueError:
        print(ValueError)


class Processes:
    def __init__(self, pid, arrival_time, burst_time):
        self.pid = pid
        self.arrival_time = arrival_time
        self.burst_time = burst_time


def sjf(processes_list, num_of_processes):
    stored_list = []
    time_unit = 1
    new_process = 0
    waiting_precesses = 0
    finished_processes = 0
    recent_pid = 0

    while finished_processes < num_of_processes:
        for i in range(new_process, num_of_processes):
            if time_unit >= processes_list[i].arrival_time:
                new_arrived_process = processes_list[i]
                stored_list.append(new_arrived_process)
                new_process += 1

        def check_new_arrival():
            global waiting_precesses
            for i in range(waiting_precesses, num_of_processes):
                if time_unit >= processes_list[i].arrival_time:
                    new_arrived = processes_list[i]
                    queue.append(new_arrived.pid)
                    d[new_arrived.pid] = 0
                    waiting_precesses += 1

        stored_list = sorted(stored_list, key=lambda x: x.burst_time)
        current_pid = stored_list[0].pid

        def check_queue():
            x = []
            for i in queue:
                d[i] += 1
                x.append(f"PID: {i} Wait={d[i]}TU")
            return ", ".join(map(str, x))

        if current_pid != recent_pid and time_unit > 1:
            print(f"Time Unit: {time_unit} Context switch. \nQueue={len(queue)} {check_queue()}")
            time_unit += 1

        if stored_list[0].burst_time > 0:
            check_new_arrival()
            queue.remove(stored_list[0].pid)
            for i in range(stored_list[0].burst_time):
                instructions = stored_list[0].burst_time - 1
                check_new_arrival()
                recent_pid = stored_list[0].pid
                if len(queue) < 1:
                    if instructions != 0:
                        print(
                            f"Time Unit: {time_unit} PID: {stored_list[0].pid} executes. "
                            f"{instructions} instructions left. No queue")
                    elif instructions == 0:
                        print(
                            f"Time Unit: {time_unit} PID: {stored_list[0].pid} "
                            f"last instruction. CPU is idle.")
                        print("All processes have executed. End of simulation.")
                else:
                    if instructions != 0:
                        print(
                            f"Time Unit: {time_unit} PID: {stored_list[0].pid} executes. "
                            f"{instructions} instructions left Queue={len(queue)} {check_queue()}")
                    elif instructions == 0:
                        print(
                            f"Time Unit: {time_unit} PID: {stored_list[0].pid} "
                            f"last instruction. Queue={len(queue)} {check_queue()}")
                time_unit += 1
                stored_list[0].burst_time -= 1
                if stored_list[0].burst_time == 0:
                    stored_list.pop(0)
                    finished_processes += 1
                    waiting_precesses -= 1


p = []
for i in processes:
    p.append(Processes(i[0], i[1], i[2]))


sjf(p, len(p))

